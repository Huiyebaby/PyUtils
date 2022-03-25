import traceback
 
import cv2
import time
import random
 
import openpyxl
import requests
import numpy as np
from PIL import Image
from io import BytesIO
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver import ActionChains
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
 
 
class CrackSlider():
    def __init__(self):
        options = webdriver.ChromeOptions()
        options.add_experimental_option('useAutomationExtension', False)
        options.add_experimental_option('excludeSwitches', ['enable-automation'])
        self.browser = webdriver.Chrome(options=options)
        self.browser.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
            "source": """
                      Object.defineProperty(navigator, 'webdriver', {
                        get: () => undefined
                      })
                    """
        })
        # self.s2 = r'//*[@id="captcha_div"]/div/div[1]/div/div[1]/img[1]'
        # self.s3 = r'//*[@id="captcha_div"]/div/div[1]/div/div[1]/img[2]'
        self.url = 'https://aq.reg.163.com/ydaq/welcome?module=offlinePasswordFind#/'  # 测试网站
 
        self.wait = WebDriverWait(self.browser, 10)
 
        self.browser.get(self.url)
 
    def refresh_url(self):
        self.browser.refresh()
 
    def get_ele_line_edit(self, num):
        ele = self.wait.until(EC.presence_of_element_located(
            (By.XPATH, '/html/body/div[1]/div/div[2]/div[1]/div[2]/div/form/div[2]/div/div[1]/div[1]/div[2]/input')))
        # time.sleep(0.5)
        ele.click()
 
        time.sleep(0.5)
        ele.send_keys(str(num))
        time.sleep(0.5)
 
    # def intput_all_num(self, num):
    #     for i in str(num):
    #         b = i
    #         key_press(str(int(b) - 1))
 
    def find_commit_button(self):
        target = self.wait.until(EC.presence_of_element_located((By.CLASS_NAME, 'bview-button')))
        target.click()
 
    def get_img(self, target, template, xp):
        time.sleep(3)
        target_link = self.wait.until(EC.presence_of_element_located((By.CLASS_NAME, 'yidun_bg-img'))).get_attribute(
            'src')
        template_link = self.wait.until(EC.presence_of_element_located((By.CLASS_NAME, 'yidun_jigsaw'))).get_attribute(
            'src')
        target_img = Image.open(BytesIO(requests.get(target_link).content))
        template_img = Image.open(BytesIO(requests.get(template_link).content))
        target_img.save(target)
        template_img.save(template)
        size_loc = target_img.size
        # print('size_loc[0]-----\n')
        # print(size_loc[0])
        zoom = xp / int(size_loc[0])  # 耦合像素
        # print('zoom-----\n')
        # print(zoom)
        return zoom
 
    def drag_success(self):
        target_link = self.wait.until(EC.presence_of_element_located((By.CLASS_NAME, 'yidun_bg-img'))).get_attribute(
            'src')
        return target_link
 
    def change_size(self, file):
        image = cv2.imread(file, 3)  # 读取图片 image_name应该是变量
        img = cv2.medianBlur(image, 9)  # 中值滤波，去除黑色边际中可能含有的噪声干扰
        b = cv2.threshold(img, 15, 255, cv2.THRESH_BINARY)  # 调整裁剪效果
        binary_image = b[1]  # 二值图--具有三通道
        binary_image = cv2.cvtColor(binary_image, cv2.COLOR_BGR2GRAY)
        x, y = binary_image.shape
        edges_x = []
        edges_y = []
        for i in range(x):
            for j in range(y):
                if binary_image[i][j] == 255:
                    edges_x.append(i)
                    edges_y.append(j)
 
        left = min(edges_x)  # 左边界
        right = max(edges_x)  # 右边界
        width = right - left  # 宽度
        bottom = min(edges_y)  # 底部
        top = max(edges_y)  # 顶部
        height = top - bottom  # 高度
        pre1_picture = image[left:left + width, bottom:bottom + height]  # 图片截取
        return pre1_picture  # 返回图片数据
 
    def match(self, target, template):
        img_gray = cv2.imread(target, 0)
        img_rgb = self.change_size(template)
        template = cv2.cvtColor(img_rgb, cv2.COLOR_BGR2GRAY)
        # cv2.imshow('template', template)
        # cv2.waitKey(0)
        res = cv2.matchTemplate(img_gray, template, cv2.TM_CCOEFF_NORMED)
        run = 1
 
        # 使用二分法查找阈值的精确值
        L = 0
        R = 1
        while run < 20:
            run += 1
            threshold = (R + L) / 2
            if threshold < 0:
                print('Error')
                return None
            loc = np.where(res >= threshold)
            if len(loc[1]) > 1:
                L += (R - L) / 2
            elif len(loc[1]) == 1:
                break
            elif len(loc[1]) < 1:
                R -= (R - L) / 2
        res = loc[1][0]
        # print('match distance-----\n')
        # print(res)
        return res
 
    def move_to_gap(self, tracks):
        slider = self.wait.until(EC.element_to_be_clickable((By.CLASS_NAME, 'yidun_slider')))
        ActionChains(self.browser).click_and_hold(slider).perform()
        # element = self.browser.find_element_by_xpath(self.s3)
        # ActionChains(self.browser).click_and_hold(on_element=element).perform()
        while tracks:
            x = tracks.pop(0)
            # print('tracks.pop(0)-----\n')
            # print(x)
            ActionChains(self.browser).move_by_offset(xoffset=x, yoffset=0).perform()
            # ActionChains(self.browser).move_to_element_with_offset(to_element=element, xoffset=x, yoffset=0).perform()
            # time.sleep(0.01)
        time.sleep(0.05)
        ActionChains(self.browser).release().perform()
 
    def move_to_gap1(self, distance):
        distance += 46
        time.sleep(1)
        element = self.browser.find_element_by_xpath(self.s3)
        ActionChains(self.browser).click_and_hold(on_element=element).perform()
        ActionChains(self.browser).move_to_element_with_offset(to_element=element, xoffset=distance,
                                                               yoffset=0).perform()
        # ActionChains(self.browser).release().perform()
        time.sleep(1.38)
        ActionChains(self.browser).release(on_element=element).perform()
 
    def move_to_gap2(self, distance):
        element = self.browser.find_elements_by_class_name("yidun_slider")[0]
        action = ActionChains(self.browser)
        mouse_action = action.click_and_hold(on_element=element)
        distance += 11
        distance = int(distance * 32 / 33)
        move_steps = int(distance / 4)
        for i in range(0, move_steps):
            mouse_action.move_by_offset(4, random.randint(-5, 5)).perform()
        time.sleep(0.1)
        mouse_action.release().perform()
 
    def get_tracks(self, distance, seconds, ease_func):
        distance += 20
        tracks = [0]
        offsets = [0]
        for t in np.arange(0.0, seconds, 0.1):
            ease = ease_func
            # print('ease-----\n')
            # print(ease)
            offset = round(ease(t / seconds) * distance)
            # print('offset-----\n')
            # print(offset)
            tracks.append(offset - offsets[-1])
            # print('offset - offsets[-1]-----\n')
            # print(offset - offsets[-1])
            offsets.append(offset)
            # print('offsets-----\n')
            # print(offsets)
        tracks.extend([-3, -2, -3, -2, -2, -2, -2, -1, -0, -1, -1, -1])
        return tracks
 
    def get_tracks1(self, distance):
        """
        根据偏移量获取移动轨迹
        :param distance: 偏移量
        :return: 移动轨迹
        """
        # 移动轨迹
        track = []
        # 当前位移
        current = 0
        # 减速阈值
        mid = distance * 4 / 5
        # 计算间隔
        t = 0.2
        # 初速度
        v = 0
 
        while current < distance:
            if current < mid:
                # 加速度为正 2
                a = 4
            else:
                # 加速度为负 3
                a = -3
            # 初速度 v0
            v0 = v
            # 当前速度 v = v0 + at
            v = v0 + a * t
            # 移动距离 x = v0t + 1/2 * a * t^2
            move = v0 * t + 1 / 2 * a * t * t
            # 当前位移
            current += move
            # 加入轨迹
            track.append(round(move))
        return track
 
    def ease_out_quart(self, x):
        res = 1 - pow(1 - x, 4)
        # print('ease_out_quart-----\n')
        # print(res)
        return res
 
    def open_url(self):
        self.browser.get(self.url)
 
 
if __name__ == '__main__':
 
    xp = 378  # 验证码的像素-长
    target = 'target.jpg'  # 临时保存的图片名
    template = 'template.png'  # 临时保存的图片名
    num_list = []
    cs = CrackSlider()
    wb = openpyxl.load_workbook('testph.xlsx')
    sheet = wb['Mysheet']   
    # print(sheet.max_row)
    i = 1
    while i <= sheet.max_row:
        # for i in range(1, sheet.max_row + 1):
        try:
            wb = openpyxl.load_workbook('testph.xlsx')
            sheet = wb['Mysheet']
            print(sheet.max_row)
            print(num_list, '当前号码')
            print(i)
            cs.open_url()
            print(sheet.cell(i, 1).value)
            num = sheet.cell(i, 1).value
            cs.get_ele_line_edit(num)
          
            time.sleep(1)
            cs.find_commit_button()
            zoom = cs.get_img(target, template, xp)
            distance = cs.match(target, template)
            track = cs.get_tracks((distance + 7) * zoom, random.randint(2, 4), cs.ease_out_quart)
            # track = cs.get_tracks1(distance)
            # track = cs.get_tracks((distance + 7) * zoom, random.randint(1, 2), cs.ease_out_quart)
            cs.move_to_gap(track)
            # cs.move_to_gap1(distance)
            # cs.move_to_gap2(distance)
            # time.sleep(1)
            try:
                # cs.browser.switch_to().fame("")
                a = cs.wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[3]/div/div/span')))
                if a.text == '帐号输入错误，请修改后重试':
                    sheet.cell(i, 2, 'F')
                    sheet.cell(i, 3, '账号不存在')
 
                    wb.save('testph.xlsx')
                    i += 1
                    continue
                else:
                    continue
            except Exception as e:
                print(e, '没有找到toast这个元素')
                msg = traceback.format_exc()
 
                now_url = cs.browser.current_url
 
                if now_url:
                    if str(now_url) == 'https://aq.reg.163.com/ydaq/welcome?module=verify#/pwdFind':
                        time.sleep(2)
                        sheet.cell(i, 2, 'T')
                        sheet.cell(i, 3, '账号存在')
                        i += 1
                        wb.save('testph.xlsx')
                        num_list.append(num)
                        cs.open_url()
                else:
                    print('拼图失败')
                    continue
        except Exception as e:
            print(e, '其他异常')
            msg = traceback.format_exc()
            continue
 
    print(num_list)
    wb.save('testph.xlsx')
    # cs.browser.close()